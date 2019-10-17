from openpyxl import load_workbook

# 加载工作簿
wb = load_workbook('./rawdata/subject.xlsx')
# 所有工作表表名 wb.sheetnames

# 使用的工作表
ws = wb[wb.sheetnames[0]]

# 最大行数
rows = ws.max_row
# 最大列数
cols = ws.max_column
tag_list = []
for i in range(2, rows + 1):
    val = ws.cell(i, 1).value
    tag_list.append(val)
print('*Vertices', len(tag_list))
for i in range(0, len(tag_list)):
    print(i + 1, '"' + tag_list[i] + '"')
# print('*Arcs')
print('*Edges')
for i in range(2, rows + 1):
    for j in range(2, cols + 1):
        val = ws.cell(i, j).value
        if j > i and val > 0:
            print(i - 1, j - 1, val)
